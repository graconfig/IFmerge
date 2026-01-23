"""模板填充模块

负责将合并后的IF数据填充到Excel模板中。
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple
from datetime import datetime
from openpyxl import load_workbook
from ebs_merger.if_grouper import IFInfo


class TemplateFiller:
    """模板填充器"""
    
    def __init__(self, template_path: str = "template/IF_Template.xlsm"):
        """初始化模板填充器
        
        参数:
            template_path: 模板文件路径
        """
        self.template_path = Path(template_path)
    
    def fill_merged_groups(
        self,
        if_dict: Dict[str, IFInfo],
        group_assignments: Dict[str, str],
        similar_pairs: List[Tuple[str, str, float]],
        input_df: pd.DataFrame,
        output_dir: str = "output",
        merged_if_names: Dict[str, str] = None
    ):
        """为所有需要合并的组填充模板
        
        参数:
            if_dict: IF信息字典
            group_assignments: IF到グルーピングIDの映射
            similar_pairs: 相似IF对列表
            input_df: 原始输入数据DataFrame
            output_dir: 输出文件夹路径
            merged_if_names: AI生成的合并IF名字典 {group_id: merged_name}
        """
        # 构建分组信息：group_id -> [if_names]
        groups = {}
        for if_name, group_id in group_assignments.items():
            if group_id not in groups:
                groups[group_id] = []
            groups[group_id].append(if_name)
        
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # 只处理需要合并的组（成员数>1）
        for group_id, group_members in groups.items():
            if len(group_members) > 1:
                # 获取AI生成的合并IF名（如果有）
                merged_name = None
                if merged_if_names and group_id in merged_if_names:
                    merged_name = merged_if_names[group_id]
                
                self.fill_single_group(
                    group_id,
                    group_members,
                    if_dict,
                    input_df,
                    output_path,
                    merged_name
                )
    
    def fill_single_group(
        self,
        group_id: str,
        group_members: List[str],
        if_dict: Dict[str, IFInfo],
        input_df: pd.DataFrame,
        output_path: Path,
        merged_if_name: str = None
    ):
        """为单个合并组填充模板
        
        参数:
            group_id: 分组ID（如G001）
            group_members: 组内IF名称列表
            if_dict: IF信息字典
            input_df: 原始输入数据DataFrame
            output_path: 输出文件夹路径
            merged_if_name: AI生成的合并IF名（可选）
        """
        # 加载模板
        wb = load_workbook(self.template_path, keep_vba=True)
        ws = wb['エクスポート項目']
        
        # 获取当前日期
        today = datetime.now().strftime('%Y/%m/%d')
        
        # 收集所有IF的文書管理番号（用第一个）
        first_if = group_members[0]
        doc_number = if_dict[first_if].doc_number
        
        # 生成合并后的文書名（IF名）
        # 如果有AI生成的名称，使用AI名称；否则使用下划线连接
        if merged_if_name:
            merged_name = merged_if_name
        else:
            merged_name = "_".join(sorted(group_members))
        
        # 安全地设置单元格值（处理合并单元格）
        def safe_set_cell(cell_ref, value):
            try:
                ws[cell_ref] = value
            except AttributeError:
                # 如果是合并单元格，找到主单元格
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        # 获取合并区域的左上角单元格
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        ws.cell(row=min_row, column=min_col).value = value
                        return
                # 如果不是合并单元格，直接设置
                ws[cell_ref].value = value
        
        # 填充基本信息
        safe_set_cell('G4', doc_number)  # 文書管理番号
        safe_set_cell('Q4', merged_name)  # 文書名（使用AI生成的名称）
        safe_set_cell('AF4', today)  # 作成日
        safe_set_cell('AQ4', today)  # 最終更新日
        safe_set_cell('G6', merged_name)  # データ定義名称（使用AI生成的名称）
        safe_set_cell('G21', f"{merged_name}.csv")  # ファイル名（使用AI生成的名称）
        
        # 收集所有IF的数据行
        all_rows = []
        for if_name in sorted(group_members):
            # 从原始数据中获取该IF的所有行
            if_rows = input_df[input_df['IF名'] == if_name]
            all_rows.append(if_rows)
        
        # 合并所有数据
        merged_data = pd.concat(all_rows, ignore_index=True)
        
        # 去重：基于EBSテーブルID和項目ID的组合去重
        merged_data = merged_data.drop_duplicates(
            subset=['EBSテーブルID', '項目ID'],
            keep='first'
        )
        
        # 重置索引
        merged_data = merged_data.reset_index(drop=True)
        
        # 填充抽出項目表格（从第28行开始）
        start_row = 28
        template_items = 10  # 模板初始项目数（每个项目占2行，所以是20行）
        row_no = 1
        
        # 复制行格式的辅助函数
        def copy_row_format(source_row, target_row):
            """复制源行的格式到目标行"""
            from copy import copy
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_row, column=col)
                target_cell = ws.cell(row=target_row, column=col)
                
                # 复制格式
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
        
        def copy_merged_cells(source_row1, source_row2, target_row1, target_row2):
            """复制源行的合并单元格到目标行
            
            参数:
                source_row1: 源起始行号
                source_row2: 源结束行号
                target_row1: 目标起始行号
                target_row2: 目标结束行号
            """
            from openpyxl.worksheet.cell_range import CellRange
            
            # 找出源行范围内的所有合并单元格
            merged_to_copy = []
            for merged_range in ws.merged_cells.ranges:
                min_row = merged_range.min_row
                max_row = merged_range.max_row
                
                # 如果合并单元格在源行范围内
                if min_row >= source_row1 and max_row <= source_row2:
                    merged_to_copy.append(merged_range)
            
            # 复制合并单元格到目标行
            for merged_range in merged_to_copy:
                # 计算行偏移
                row_offset = target_row1 - source_row1
                
                # 创建新的合并单元格范围
                new_min_row = merged_range.min_row + row_offset
                new_max_row = merged_range.max_row + row_offset
                new_min_col = merged_range.min_col
                new_max_col = merged_range.max_col
                
                # 合并目标单元格
                ws.merge_cells(
                    start_row=new_min_row,
                    start_column=new_min_col,
                    end_row=new_max_row,
                    end_column=new_max_col
                )
        
        for idx, row in merged_data.iterrows():
            current_row = start_row + (row_no - 1) * 2
            
            # 如果超过模板初始行数，需要复制格式和合并单元格
            if row_no > template_items:  # 前10个项目（20行）是模板自带的
                # 复制第10个项目的格式（第46-47行）作为模板
                template_row1 = start_row + (template_items - 1) * 2  # 第10个项目的第一行（第46行）
                template_row2 = start_row + (template_items - 1) * 2 + 1  # 第10个项目的第二行（第47行）
                
                copy_row_format(template_row1, current_row)
                copy_row_format(template_row2, current_row + 1)
                
                # 复制合并单元格
                copy_merged_cells(template_row1, template_row2, current_row, current_row + 1)
            
            # B列：No（流水号）
            safe_set_cell(f'B{current_row}', row_no)
            
            # C列：テーブル名（格式：表名 (表ID)）
            table_name = row['EBSテーブル名']
            table_id = row['EBSテーブルID']
            
            # 第一行显示：表名 (表ID)
            if pd.notna(table_name) and pd.notna(table_id):
                table_display = f"{table_name} ({table_id})"
            elif pd.notna(table_name):
                table_display = str(table_name)
            else:
                table_display = ""
            
            safe_set_cell(f'C{current_row}', table_display)
            # 第二行留空或可以不设置
            
            # R列：項目（第一行：項目ID，第二行：項目名）
            item_id = row['項目ID']
            item_name = row['項目名']
            safe_set_cell(f'R{current_row}', str(item_id) if pd.notna(item_id) else "")
            safe_set_cell(f'R{current_row + 1}', str(item_name) if pd.notna(item_name) else "")
            
            row_no += 1
        
        # 保存文件（文件名也使用AI生成的名称）
        output_filename = f"{group_id}_{merged_name}.xlsm"
        output_filepath = output_path / output_filename
        wb.save(output_filepath)
        
        print(f"  テンプレートファイルを生成しました：{output_filename}")
